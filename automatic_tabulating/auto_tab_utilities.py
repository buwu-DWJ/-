#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   auto_tab_utilities.py
@Time    :   2023/01/09 10:05:18
@Author  :   DingWenjie
@Contact :   359582058@qq.com
@Desc    :   None
'''


import os
import ctypes
import sys
import warnings
import time
import datetime
import smtplib
import psutil
import pyautogui
import tkinter as tk
import pandas as pd
import openpyxl as op
import numpy as np
from tkinter import ttk
from icetcore import TCoreAPI, OrderStruct
from smtplib import SMTP_SSL
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
from email.mime.application import MIMEApplication


warnings.simplefilter('ignore')
thisDir = os.path.dirname(__file__)
api = TCoreAPI()
re = api.connect()
tradingday = pd.read_hdf(os.path.join(thisDir, 'tradingday.h5'))
date_str_list = [csd_date.strftime('%Y%m%d') for csd_date in tradingday.index]
TODAY_STR = datetime.date.today().strftime('%Y%m%d')
YESTERDAY_STR = date_str_list[date_str_list.index(TODAY_STR)-1]


class Auto_Tab():
    def __init__(self, account_list, und_list, today_str=TODAY_STR, yesterday=YESTERDAY_STR):
        self.account_list = account_list
        self.und_list = und_list
        self.today_str = today_str
        self.yesterday_str = yesterday
        workbook = op.load_workbook(thisDir+'\\每日损益.xlsx')
        worksheet = workbook['总览']
        print(worksheet.cell(row=1, column=1).value)

    def __find_files_given_date(self, date_str):
        inputFilePath = os.path.join(thisDir, 'data')
        file_list = []
        for account in self.account_list:
            key_0 = date_str + '期权持仓报表' + account
            key_1 = date_str + '期权成交报表' + account
            for file in os.listdir(inputFilePath):
                if os.path.splitext(file)[1] == '.csv':
                    if (key_0 in file and file.split(key_0)[1][0] == '.') or (key_1 in file and file.split(key_1)[1][0] == '.'):
                        sourcefile = os.path.join(inputFilePath, file)  # 拼路径
                        file_list += [sourcefile]
        return file_list

    @staticmethod
    def __convert_date_str(date_str):
        '''convert date str between formated '%YYYY%MM%DD' and '%YYYY年%MM月%%DD日' '''
        converted_date_str = date_str[:4]+'年' + \
            str(int(date_str[4:6]))+'月'+str(int(date_str)[6:])+'日'
        return converted_date_str

    def convert_excel_to_hdf(self):
        today_str_ = self.__convert_date_str(self.today_str)
        yesterday_str_ = self.__convert_date_str(self.yesterday_str)
        today_file_list = self.__find_files_given_date(today_str_)
        yesterday_file_list = self.__find_files_given_date(yesterday_str_)
        total_list = today_file_list+yesterday_file_list

        return

    @staticmethod
    def convert_option_code(code):
        '''将汇点获取的合约代码转换为权分析中的合约代码'''
        converted_code = api.symbollookup(findkey=code, symboltype='OPT')
        if len(converted_code) != 1:
            sys.exit(f'合约{code}的模糊查找有错误!')
        return converted_code

    @staticmethod
    def get_position_of_different_und_given_account(account):
        '''给定账户, '''

    def get_total_position_given_account_and_und(self, account, und):
        pass

    def get_intraday_position_given_account_and_und(self, account, und):
        pass

    def get_yesterday_position_given_today_and_intraday_position(self, today_position, intraday_position):
        pass

    def get_greeks_given_position(self, position):
        pass

    def write_position_and_greeks_given_account_and_und(self, account, und):
        ''''''

    def process_data(self):
        pass

    def print_excel(self):
        self.capslock()
        try:
            op.load_workbook(os.path.join(
                thisDir, f'data\每日损益{self.today_str}.xlsx'))
            op.load_workbook(os.path.join(
                thisDir, f'data\每日损益{self.today_str}_no_ls7.xlsx'))
        except:
            print('未找到今日损益excel')
        for file_str in ['', '_no_ls7']:
            file = os.path.join(
                thisDir, f'data\每日损益{self.today_str}{file_str}.xlsx')
            self.open_excel_and_full_screen(file)
            # 打开Acrobat加载项
            pyautogui.moveTo(829, 47)
            time.sleep(1)
            pyautogui.click()
            time.sleep(1)
            # 创建pdf
            pyautogui.moveTo(25, 93)
            time.sleep(1)
            pyautogui.click()
            time.sleep(1)
            # 转换范围: 整个工作簿
            pyautogui.moveTo(776, 340)
            time.sleep(1)
            pyautogui.click()
            time.sleep(1)
            # 转换选项: 适合纸张宽度
            pyautogui.moveTo(777, 673)
            time.sleep(1)
            pyautogui.click()
            time.sleep(1)
            # 转换为pdf
            pyautogui.hotkey('c')
            time.sleep(1)
            # 覆盖
            pyautogui.hotkey('y')
            time.sleep(5)

    @staticmethod
    def capslock():
        hllDll = ctypes.WinDLL("User32.dll")
        VK_CAPITAL = 0x14
        if ((hllDll.GetKeyState(VK_CAPITAL)) & 0xffff) != 0:
            return
        else:
            pyautogui.hotkey('capslock')
            time.sleep(1)
            return

    @staticmethod
    def open_excel_and_full_screen(file):
        os.startfile(file)
        time.sleep(5)
        pyautogui.hotkey('win', 'up')
        print(1)
        time.sleep(2)

    def send_mail(self):
        host_server = 'smtp.163.com'  # qq邮箱smtp服务器
        sender_163 = '13918949838@163.com'  # 发件人邮箱
        pwd = 'EGGZASTFLHVGCBRU'
        mail_title = f'每日损益{self.today_str}'
        mail_content = ""
        receiver = '359582058@qq.com'
        msg = MIMEMultipart()
        msg["Subject"] = Header(mail_title, 'utf-8')
        msg["From"] = sender_163
        msg["To"] = Header("测试邮箱", "utf-8")
        msg.attach(MIMEText(mail_content, 'html'))
        basename = f'每日损益{self.today_str}.pdf'
        attachment = MIMEApplication(
            open(os.path.join(thisDir, f'data\每日损益{self.today_str}.pdf'), 'rb').read())
        attachment["Content-Type"] = 'application/octet-stream'
        attachment.add_header('Content-Disposition', 'attachment',
                              filename=('utf-8', '', basename))
        msg.attach(attachment)
        try:
            smtp = SMTP_SSL(host_server)  # ssl登录连接到邮件服务器
            smtp.set_debuglevel(1)  # 0是关闭，1是开启debug
            smtp.ehlo(host_server)  # 跟服务器打招呼，告诉它我们准备连接，最好加上这行代码
            smtp.login(sender_163, pwd)
            smtp.sendmail(sender_163, receiver, msg.as_string())
            smtp.quit()
            print("邮件发送成功")
            return True
        except smtplib.SMTPException:
            print("无法发送邮件")
            return False

    def auto_tab_gui(self):
        window = tk.Tk()
        width = 280
        height = 150
        screenwidth = window.winfo_screenwidth()
        screenheight = window.winfo_screenheight()
        size_geo = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth-width)/2, (screenheight-height)/2)
        window.geometry(size_geo)
        window.title('期权损益制表')
        pids = psutil.pids()
        t_on = 0
        for pid in pids:
            p = psutil.Process(pid)
            if "星驰期权客户端" in p.name():
                t_on = 1
        if t_on:
            tk.Label(window, text='系统检测到星驰运行中, 可进行下一步').grid(
                row=0, columnspan=2)
        else:
            tk.Label(window, text='系统检测到星驰未运行, 请开启星驰!').grid(
                row=0, columnspan=2)
        tk.Button(window, text="处理数据", width=10, command=self.process_data).grid(
            row=1, column=0, padx=1, pady=1)
        tk.Button(window, text="生成pdf", width=10, command=self.print_excel).grid(
            row=2, column=0, padx=1, pady=1)
        tk.Button(window, text="发送邮件", width=10, command=self.send_mail).grid(
            row=3, column=0, padx=1, pady=1)
        do_sending = ttk.Combobox(
            window, textvariable=tk.StringVar(), width=10)
        do_sending['values'] = ('确认发送', '暂不发送')
        do_sending.current(1)
        # do_sending.bind('<<ComboboxSelected>>', go_0)
        do_sending.grid(row=3, column=1)
        window.mainloop()
