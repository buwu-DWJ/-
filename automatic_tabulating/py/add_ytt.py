# -*- coding: utf-8 -*-
"""
Created on Wed Oct 19 16:19:09 2022

@author: mrlia
"""

import openpyxl as op
import time
import pandas as pd
from WindPy import w
import os
import string
w.start()
w.menu()


def find_account_files(acc, month):
    inputFilePath= "D:\wind-trial"
    file_list = []
    key = acc + '每日损益' + month
    for file in os.listdir(inputFilePath):
        if os.path.splitext(file)[1]=='.xlsx':
            if key in file and file.split(key)[0]=='':
        #        if string.find(file, key) != -1:  # 满足条件往下进行
                sourcefile = os.path.join(inputFilePath, file)  # 拼路径
                file_list += [sourcefile]
    return file_list


tradingday = list(pd.read_excel('TradeDay.xls')['date'])
tradingday = [str(a)[:4]+str(a)[5:7]+str(a)[8:10] for a in tradingday]
today_str = time.strftime('%Y%m%d',time.localtime(time.time()))
today_idx = tradingday.index(today_str)
yes_str = tradingday[today_idx-1]
crt_month = str(time.localtime().tm_year)+str(time.localtime().tm_mon)
sheet_dict = {
        '德睿': ['DB', 'LS'],
        'LS7': ['LS7'],
        '聊塑3号、4号': ['LS3', 'LS4']
        }
code_dict = {
        'ETF500': '510500.SH',
        'ETF50': '510050.SH',
        'SH300': '510300.SH',
        'ETF159915': '159915.SZ'
        }
rtn_dict = {}
for i, csd_code in enumerate(['ETF50','SH300','ETF500','ETF159915']):
    _, close = w.wsd(code_dict[csd_code],"close","ED-1TD",today_str[:4]+'-'+today_str[4:6]+'-'+today_str[6:],usedf=True)
    rtn_dict[csd_code] = close['CLOSE'][-1]/close['CLOSE'][-2]-1
    print(f'{csd_code}涨跌幅为:{rtn_dict[csd_code]}')


for i, csd_account in enumerate(['德睿','LS7','聊塑3号、4号']):
    wb_today = op.load_workbook(csd_account+'每日损益'+today_str+'.xlsx')
    # 昨日对今日影响
    wb_yes = op.load_workbook(csd_account+'每日损益'+yes_str+'.xlsx')
    for j, csd_sheet in enumerate(['ETF50','SH300','ETF500','ETF159915']):
        for k, csd_subacc in enumerate(sheet_dict[csd_account]):
            ws_today = wb_today[csd_sheet+csd_subacc+'Greeks']
            ws_yes = wb_yes[csd_sheet+csd_subacc+'Greeks']
            ws_today.cell(row=20, column=1).value = '昨日delta交易影响'
            ws_today.cell(row=21, column=1).value = round(ws_yes.cell(row=15, column=2).value*rtn_dict[csd_sheet],1)
            ws_today.cell(row=22, column=1).value = '去除delta后昨日交易影响'
            ws_today.cell(row=23, column=1).value = ws_today.cell(row=18, column=5).value - ws_today.cell(row=21, column=1).value
    # 本月总体昨日交易影响
    file_list = find_account_files(csd_account, crt_month)
    for j in range(4):
        for k, csd_subacc in enumerate(sheet_dict[csd_account]):
            locals()[f'ytt_{j}_{k}'] = 0
    for file in file_list:
        for j, csd_sheet in enumerate(['ETF50','SH300','ETF500','ETF159915']):
            for k, csd_subacc in enumerate(sheet_dict[csd_account]):
                ws_today = wb_today[csd_sheet+csd_subacc+'Greeks']
                locals()[f'ytt_{j}_{k}'] += ws_today.cell(row=23, column=1).value
    
    for j, csd_sheet in enumerate(['ETF50','SH300','ETF500','ETF159915']):
        for k, csd_subacc in enumerate(sheet_dict[csd_account]):
            try:
                ws_today = wb_today['损益情况']
            except:
                ws_today = wb_today[f'{csd_subacc}损益情况']
            ws_today.cell(row=33+j, column=1).value = csd_sheet
            ws_today.cell(row=32, column=2+k).value = csd_subacc
            ws_today.cell(row=33+j, column=2+k).value = locals()[f'ytt_{j}_{k}']
    wb_today.save(csd_account+'每日损益'+today_str+'.xlsx')
