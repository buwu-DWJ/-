#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   get_all_iv.py
@Time    :   2022/11/29 09:34:48
@Author  :   DingWenjie
@Contact :   359582058@qq.com
@Desc    :   获取所有品种期权 iv, skew近一周变化
'''


from PIL import ImageGrab
import xlwings as xw
import os
import openpyxl
import sys
import pandas as pd
import numpy as np
import warnings
import time
import datetime
import tcoreapi_mq as t


warnings.simplefilter('ignore')
thisDir = os.path.dirname(__file__)
core = t.TCoreZMQ(quote_port="51864", trade_port="51834")
md_str = ''
tradingday = pd.read_excel(r'C:\Users\dingwenjie\Desktop\demo\备份\-\automatic_trading\vanna_daily_20220808\tradingday.xlsx')['date']
tradingday_list = [a.strftime('%Y%m%d') for a in tradingday]
tradingday_int = np.array([int(a) for a in tradingday_list])
today = datetime.datetime.today().strftime('%Y%m%d')
try:
    today_idx = tradingday_list.index(today)
except ValueError:
    today_int = int(today)
    end_idx = (tradingday_int<today_int).sum()-1
    end_date = tradingday_list[end_idx]
else:
    if time.localtime().tm_hour>14:
        end_idx = today_idx
    else:
        end_idx = today_idx - 1
    end_date = tradingday_list[end_idx]
start_idx = end_idx - 5
start_date = tradingday_list[start_idx]
print(f'考虑的最新日期为{end_date}, 上个日期为{start_date}')
df = pd.DataFrame()
echg_ = []
und = []
iv_old = []
iv_new = []
iv_diff = []
skew_old = []
skew_new = []
skew_diff = []
a = core.QueryAllInstrumentInfo('Options')
if a['Success']!='OK':
    sys.quit('获取合约信息失败, 请重试!')
echg_list = a['Instruments']['Node']
for csd_echg in echg_list:
    print('\n当前计算'+csd_echg['CHS'])
    try:
        csd_echg['Node']
    except KeyError:
        print(' '*4+'当前交易所无期权')
        continue
    for csd_option in csd_echg['Node']:
        if csd_echg['ENG'] not in echg_:
            echg_ += [csd_echg['ENG'][:-3]]
        else:
            echg_ += ['']
        sample_option = core.QueryInstrumentInfo(csd_option['Node'][0]['Node'][0]['Contracts'][0])
        while sample_option['Success']!='OK':
            sample_option = core.QueryInstrumentInfo(csd_option['Node'][0]['Node'][0]['Contracts'][0])
        for objs in sample_option['Info']:
            try:
                und_f = sample_option['Info'][objs]['Underlying.F']+'.HOT'
                break
            except:
                pass
        dogsk_df = pd.DataFrame(core.SubHistory(und_f, 'DOGSK', start_date+'00', end_date+'07'))
        try:
            iv_array = pd.to_numeric(dogsk_df['iv']).values
        except KeyError:
            hot_month = csd_option['Node'][0]['Node'][0]['Contracts'][0].split('.')[4]
            und_f = sample_option['Info'][objs]['Underlying.F']+'.'+hot_month
            dogsk_df = pd.DataFrame(core.SubHistory(und_f, 'DOGSK', start_date+'00', end_date+'07'))
            iv_array = pd.to_numeric(dogsk_df['iv']).values
        print(' '*4+csd_option['CHS']+': '+und_f)
        if ')'in csd_option['CHS']:
            und += [csd_option['CHS'][:-10]]
        else:
            und += [csd_option['CHS'][:-2]]
        t_array = pd.to_numeric(dogsk_df['t']).values
        csd_idx = dogsk_df.index[t_array==65500000]
        cskew_array = pd.to_numeric(dogsk_df['cskew']).values
        pskew_array = pd.to_numeric(dogsk_df['pskew']).values
        skew_array = (cskew_array-pskew_array) * 10000
        iv_old += [iv_array[csd_idx[0]]]
        iv_new += [iv_array[csd_idx[-1]]]
        iv_diff += [iv_array[csd_idx[-1]]-iv_array[csd_idx[0]]]
        skew_old += [skew_array[csd_idx[0]]]
        skew_new += [skew_array[csd_idx[-1]]]
        skew_diff += [skew_array[csd_idx[-1]]-skew_array[csd_idx[0]]]
df['交易所'] = echg_
df['品种'] = und
df['iv_5d_'] = iv_old
df['iv_now'] = iv_new
df['iv_diff'] = iv_diff
df['skew_5d_'] = skew_old
df['skew_now'] = skew_new
df['skew_diff'] = skew_diff
df.index = df['交易所']
del df['交易所']
md_str += '# 全品种期权iv与skew本周变化\n' + '最新日期: '+ end_date + ', 老日期: ' + start_date + '\n' + df.to_markdown()
output_path = os.path.join(thisDir, f'summary\\{end_date}.md')
folder_path = os.path.join(thisDir, 'summary')
if not os.path.exists(folder_path):
    os.mkdir(folder_path)
if os.path.exists(output_path):
    os.remove(output_path)
with open(output_path, 'w', encoding='utf-8') as file:
    file.writelines(md_str)
wb=openpyxl.load_workbook("iv_skew_compare.xlsx")
ws = wb['Sheet1']
for row in range(len(df)):
    csd_row = row+2
    aa = ws.cell(csd_row, 1, value=df.index[row])
    for j in range(7):
        aa = ws.cell(csd_row, j+2, value=df.iloc[row, j])
ws.append(['开始日期:',start_date,'结束日期:',end_date])
wb.save('iv_skew_compare.xlsx')
app=xw.App(visible=True,add_book=False)
wb = app.books.open('iv_skew_compare.xlsx')
sheet=wb.sheets[0]
all=sheet.used_range
all.api.CopyPicture()
sheet.api.Paste()
img_name='iv_skew_compare'
pic=sheet.pictures[0]
pic.api.Copy()
img = ImageGrab.grabclipboard()
img.save(img_name + ".png")
pic.delete()
wb.close()
